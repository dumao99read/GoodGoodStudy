"""
进程池的详细教程：
https://blog.csdn.net/chusheng1840/article/details/142736660?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522a7ec64e7347c41f880190c6dca690a7c%2522%252C%2522scm%2522%253A%252220140713.130102334..%2522%257D&request_id=a7ec64e7347c41f880190c6dca690a7c&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~sobaiduend~default-1-142736660-null-null.142^v100^pc_search_result_base9&utm_term=multiprocessing%E8%BF%9B%E7%A8%8B%E6%B1%A0&spm=1018.2226.3001.4187
"""


import os
import time
import pandas as pd
import openpyxl
import multiprocessing
import logging
import win32com.client as win32

# TODO: 这个导入，会导致本文件的main里面的执行日志变成了file_setting.py里面的，待定位

# from tools import file_setting

CURRENT_DIR = os.getcwd()

logger = logging.getLogger()
logging.basicConfig(level=logging.INFO,
                    encoding='utf-8',
                    filename='stripSpace.log',
                    filemode='w',  # 模式，a为追加，w为覆盖写
                    format='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s : %(message)s')


def get_fileList_by_directory(path, directory, file_types = ['.xlsx','.xlsm']):
    file_path = os.path.join(path, directory)
    file_list = []
    for dirpath, dirnames, filenames in os.walk(file_path):
        # 获取路径下所有文件夹
        # for item in dirnames:
        #     file_list.append(os.path.join(dirpath, item))

        # 获取路径下所有指定类型的文件
        for file in filenames:
            if any(file.lower().endswith(file_type) for file_type in file_types):
                file_list.append(os.path.join(dirpath,file))

    return file_list

def check_process(i, file, log_path):
    file_name = os.path.split(file)[1]
    log_file = os.path.join(log_path, f'{file_name}-日志.log')

    process_1 = time.time()
    logging.info(f'{i}号子进程开始：{file}')
    print(f'{i}号子进程开始：{file}')
    logger.warning(f'当前检查文件：{file}')

    wb = openpyxl.load_workbook(file)
    sheet_name_list = wb.sheetnames

    # 开启记录，存放检查日志
    handle = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    handle.setLevel(logging.WARNING)
    logger.addHandler(handle)

    for sheet_name in sheet_name_list:
        data = pd.read_excel(file, sheet_name=sheet_name, header=None)  # 注释测试2
        for row_index, row in data.iterrows():
            for col_index, cell_value in enumerate(row):
                if str(cell_value).startswith(' ') and str(cell_value).endswith(' '):
                    logger.warning(f'{sheet_name}页签第{row_index + 1}行{chr(col_index + 65)}列：已检查到有前后空格[{cell_value}]')
                elif str(cell_value).startswith(' '):
                    logger.warning(f'{sheet_name}页签第{row_index + 1}行{chr(col_index + 65)}列：已检查到有前空格[{cell_value}]')
                elif str(cell_value).endswith(' '):
                    logger.warning(f'{sheet_name}页签第{row_index + 1}行{chr(col_index + 65)}列：已检查到有后空格[{cell_value}]')

    process_2 = time.time()
    logging.info(f'{i}号子进程结束：{file}')
    logging.info(f'{i}号子进程耗时为：{process_2 - process_1}')
    print(f'{i}号子进程结束：{file}')
    print(f'{i}号子进程耗时为：{process_2 - process_1}')

def calculate_excel(file):
    # 启动Excel应用程序
    excel = win32.Dispatch("Excel.Application")
    # 设置为False可以避免Excel窗口弹出
    excel.visible = False
    # 打开文件
    workbook = excel.Workbooks.Open(file)
    # 计算所有公式
    excel.Calculate()
    # 保存文件
    workbook.Save()
    # 关闭Excel应用程序
    workbook.Close()
    excel.Quit()

def strip_process(file,strip_char=''):
    if file.lower().endswith('.xlsm'):
        wb = openpyxl.load_workbook(file, keep_vba=True)
    else:
        wb = openpyxl.load_workbook(file)
    sheet_name_list = wb.sheetnames

    for sheet_name in sheet_name_list:
        ws = wb[sheet_name]
        data = pd.read_excel(file, sheet_name=sheet_name, header=None)
        if not strip_char:
            # 默认去空格
            data = data.map(lambda x: x.strip() if isinstance(x, str) else x)
        else:
            data = data.map(lambda x: x.strip(strip_char) if isinstance(x, str) else x)
        for row_index, row in data.iterrows():
            for col_index, cell_value in enumerate(row):
                cell  =ws.cell(row=row_index + 1, column=col_index + 1)
                if pd.isna(cell_value):
                    continue  # pandas的合并单元格，子单元格是已NaN的形式出现的，遇到就跳过
                # 判断单元格是否包含公式，如果包含公式，则cell_value取源公式
                if cell.data_type == 'f':
                    cell_value = cell.value
                else:
                    ws.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
    wb.save(file)

    # 调用以下方法能够保证公式被正确计算，然后pandas读取公式就不会出现NaN的问题
    calculate_excel(file)


def check_space(path, directory):
    time1 = time.time()
    log_path = os.path.join(CURRENT_DIR, '../checkSpaceLog')
    if not os.path.isdir(log_path):
        os.mkdir(log_path)

    file_list = get_fileList_by_directory(path, directory)

    cpu_count = os.cpu_count()
    process_num = int(cpu_count / 2)

    # pool.apply()方法，只能同步执行进程，不能异步执行。

    # 如果进程方法是单个传参，使用map，可将可迭代的列表file_list中的元素传给进程方法:pool.map(check_process, file_list)
    # 如果进程方法是多个传参，使用starmap，可将列表中的元祖元素传给进程方法:pool.starmap(check_process, [(1,2),(3,4)])
    pool = multiprocessing.Pool(processes=process_num)

    for i, file in enumerate(file_list):
        pool.apply_async(check_process, args=(i, file, log_path))

    pool.close()
    pool.join()

    time2 = time.time()
    print('检查空格整体耗时为：{}'.format(time2 - time1))

def strip_space(path, directory):
    time1 = time.time()
    file_list = get_fileList_by_directory(path, directory)

    cpu_count = os.cpu_count()
    process_num = int(cpu_count / 2)
    pool = multiprocessing.Pool(processes=process_num)

    pool.map(strip_process, file_list)

    time2 = time.time()
    print('去除空格整体耗时为：{}'.format(time2 - time1))



if __name__ == '__main__':
    check_space(CURRENT_DIR, '../fileDemo')
    # strip_space(CURRENT_DIR, '../fileDemo')
    # file_setting.format_excel_by_openpyxl(os.path.join(CURRENT_DIR, '../fileDemo/测试3.xlsx'))

