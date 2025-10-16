import datetime
import logging
import os
import time

import pytz
import xlwings as xw
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from common.tryExcept import try_except

# logging.basicConfig(level=logging.INFO,
#                     encoding='utf-8',
#                     filename='file_seeting.log',
#                     filemode='w',  # 模式，a为追加，w为覆盖写
#                     format='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s : %(message)s')

def format_excel_by_xlwings(file_path, format_list=[]):
    app = xw.App(visible=False, add_book=False)  # 打开Excel工具
    wb = app.books.open(file_path)  # 打开指定工作簿
    sheets = wb.sheets

    # 定义边框线条
    line_style_value = 1  # 1是直线，2是虚线
    weight = 2  # 2是细线，3是粗线

    try:
        for sheet in sheets:
            # 获取当前工作表的总行数
            rows = sheet.used_range.last_cell.row
            # 获取当前工作表的总列数
            cols = sheet.used_range.last_cell.column
            # 获取所有单元格
            cell = sheet[:rows, :cols]

            sheet[0, 0:cols].color = (197, 217, 241)  # 设置标题背景颜色格式
            sheet[0, 0:cols].api.Font.Name = '微软雅黑'

            # 列遍历，根据format_list中传递的列宽进行调整
            for i,j in enumerate(format_list):
                sheet[0, i].column.width = j

            # 设置所有单元格格式
            cell.api.VerticalAlignment = -4130  # 上下对齐方式：-4108：垂直居中(默认)；-4160：顶端对齐；-4107：底端对齐；-4130：自动换行对齐
            # Borders(7)到Borders(10)：上下左右框线； Borders(11)：垂直框线； Borders(12)：水平框线
            cell.api.Borders(7).LineStyle = line_style_value
            cell.api.Borders(7).Weight = weight
            cell.api.Borders(8).LineStyle = line_style_value
            cell.api.Borders(8).Weight = weight
            cell.api.Borders(9).LineStyle = line_style_value
            cell.api.Borders(9).Weight = weight
            cell.api.Borders(10).LineStyle = line_style_value
            cell.api.Borders(10).Weight = weight
            cell.api.Borders(11).LineStyle = line_style_value
            cell.api.Borders(11).Weight = weight
            cell.api.Borders(12).LineStyle = line_style_value
            cell.api.Borders(12).Weight = weight
    except:  # 异常退出不保存
        wb.close()
        app.kill()  # 强制退出工具
        print('处理异常，请检查代码！')

    wb.save()
    wb.close()
    app.kill()
    print('美化结束，请查看文件！\n{}'.format(file_path))
    return file_path

@try_except
def format_excel_by_openpyxl(file_path, format_list=[], filter_type=0, freez_cell=''):
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.worksheets
    try:
        for sheet in sheets:
            # 获取当前工作表的总行数
            rows = sheet.max_row
            # 获取当前工作表的总列数
            cols = sheet.max_column
            # 获取所有单元格
            cell = sheet.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols)

            # 设置所有单元格的格式，包括列头
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                                    fill_type='solid')  # 设置标题背景颜色格式
            cell.font = Font(name='微软雅黑')
            cell.alignment = Alignment(vertical='center')  # 上下对齐方式：垂直居中(默认)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                 left=Side(style='thin'), right=Side(style='thin'))  # 设置单元格边框

            # 单独设置列头
            for item in range(0, cols):
                header_cell = sheet.cell(row=1, column=item + 1)
                header_cell.font = Font(name='微软雅黑')
                header_cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
                header_cell.alignment = Alignment(wrap_text=True)  # 列头自动换行

            # 列遍历，根据format_list中的列宽进行调整
            for i, j in enumerate(format_list):
                sheet.colunm_dimensions[get_column_letter(i + 1)].width = j

            # 列头打开自动筛选
            if filter_type == 1:
                sheet.auto_filter.ref = sheet.dimensions

            # 冻结单元格，如果要冻结首航，则为'A2'单元格
            if freez_cell:
                sheet.freeze_panes = freez_cell
        wb.save(file_path)
    except Exception as e:
        print('处理异常，请检查代码！\n{}'.format(e))

    print('美化结束，请查看文件！\n{}'.format(file_path))
    return file_path

def get_project_root():
    """根据requirements.txt文件寻找项目根目录"""
    cwd = os.getcwd()
    while True:
        if os.path.exists(os.path.join(cwd, 'requirements.txt')):
            return cwd
        parent_dir = os.path.dirname(cwd)
        if parent_dir == cwd:
            return None
        cwd = parent_dir

def get_current_time_stamp(format_type='%Y%m%d%H%M%S'):
    """根据当前系统时区来获取指定格式的时间戳(当地时间)"""
    local_time_zone = datetime.datetime.utcnow().astimezone().tzinfo
    time_stamp =  datetime.datetime.now(tz=local_time_zone).strftime(format_type)
    # time_stamp = datetime.datetime.now(tz=pytz.timezone('Asia/Shanghai')).strftime(format_type)
    return time_stamp

def wait_and_loading(msg, dp=5, loop=2):
    """
    等待加载效果，默认5*2=10秒，适用于异步任务轮询时，控制台不想输出太多信息。
    :param msg: 要补充的提示信息
    :param dp: 小数点的数量
    :param loop: 循环加载的次数
    """
    print(msg, end='')
    i = 0
    while i < loop:
        for item in range(dp):
            print('.', end='')
            time.sleep(1)
        print('\b' * dp, end='')  # 清空小数点
        i += 1
    print('\b' * 999, end='')  # 清空屏幕


class NumberIterator:
    def __init__(self):
        self.current = 1  # 初始值为1

    def __iter__(self):
        return self  # 返回迭代器对象本身

    def __next__(self):
        if self.current > 10:
            raise StopIteration  # 超过指定值时停止迭代
        num = self.current
        self.current += 1  # 数值加1，供下次调用
        return num

if __name__ == '__main__':
    # x = get_current_time_stamp()
    # print(x)

    # 轮询等待，以迭代器计数来模拟异步任务的完成状态
    x = NumberIterator()
    time_start = time.time()
    for i in range(100):
        y = next(x)
        print(y)  # 实际使用中不希望一行行打印y的值，就可以注释掉此行
        if y == 4:
            time_end = time.time()
            print(f'任务已完成，耗时为：{round(time_end - time_start, 3)}秒。')
            break
        else:
            wait_and_loading('正在遍历迭代器，请稍等', dp=3, loop=1)


